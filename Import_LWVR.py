"""
Import_LWVR
by Joseph Arnson
on Tuesday, July 21, 2020 at 02:52

The script takes the Load Weight Variance report and
parses it down to just the Shipments that need to
be actioned on. Then it removes the remaining excess
columns, formats and calculates variables, then writes
the data frame into a dictionary that will be used 
later for default Unplanned Demand assignment.

Run Time 1:     0:00:19.556800
Run Time 2:     0:00:19.721495
Run Time 3:     0:00:18.895732
Run Time 4:     0:00:18.976302
Run Time 5:     0:00:19.624358

Avg Time:       0:00:19.354937
"""

# import packages
import os
import datetime
import pandas as pd

# import functions
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

# print message after packages imported successfully
print("Import of packages successful")

# set working directory
os.chdir('C:\\Users\\JSVAR\\OneDrive\\Desktop\\Light Loads Template\\LL Python')
print(f"Directory set to {os.getcwd()}")

# start timer
begin_time = datetime.datetime.now()
print(f"Start time: {begin_time}")

"""Get source data and create LWVR Dictionary"""
# establish variables
Dunnage = 2000
DeLength = 9

# open source report
LWVR_Source = 'Load Weight Variance Report_Sample.xlsx'
wb1 = load_workbook(LWVR_Source)
ws1 = wb1["All"]

# create destination workbook
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Sheet1"

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)
        # writing the read value to destination excel file
        cell = ws2.cell(row=i, column=j)
        cell.value = c.value

# format columns 1 and 2 as numbers
for i in range(2, mr + 1):
    for j in range(1, 3):
        # ignore blank values
        try:
            c = ws2.cell(row=i, column=j)
            value = int(c.value)
            c.value = value
        except TypeError as e:
            continue

# delete unwanted columns
LWVR = wb2['Sheet1']
LWVR.delete_cols(7, 1)
LWVR.delete_cols(8, 3)
LWVR.delete_cols(10, 1)
LWVR.delete_cols(14, 1)

# insert calculated fields
ws2['S1'] = "To Fill"
ws2['T1'] = "Include"

# get new max rows and columns
mr = ws2.max_row
mc = ws2.max_column

# calculate To Fill values
for i in range(2, mr + 1):
    # ignore blank values
    try:
        c1 = ws2.cell(row=i, column=17)  # Total Weight
        c2 = ws2.cell(row=i, column=14)  # Resource Max Weight
        d = ws2.cell(row=i, column=19)  # To Fill
        # To Fill = Resource Max Weight - Dunnage - Total Weight
        d.value = c2.value - Dunnage - c1.value
    except TypeError as e:
        d = ws2.cell(row=i, column=19)  # To Fill
        d.value = -1
        continue

"""
Conditions to NOT include Delivery
1. Length Delivery < 9 (DeLength)
2. Max Resource Weight Status = Tentative
3. Multiple Deliveries on Shipment = X
4. Distributed to EWM = X
"""

for i in range(2, mr + 1):
    c1 = ws2.cell(row=i, column=2)  # Delivery
    c2 = ws2.cell(row=i, column=15)  # Resource Max Weight Status
    c3 = ws2.cell(row=i, column=12)  # Multiple Deliveries on the Shipment
    c4 = ws2.cell(row=i, column=11)  # Distributed to EWM
    c5 = ws2.cell(row=i, column=19)  # To Fill
    d = ws2.cell(row=i, column=20)  # Include
    if len(str(c1.value)) < DeLength or c2.value == 'Tentative' or c3.value == 'X' \
            or c4.value == 'X' or c5.value <= 0:
        d.value = 0
    else:
        d.value = 1

# saving the destination excel file
wb2.save('LWVR.xlsx')

# convert xlsx to csv and remove xlsx
read_file = pd.read_excel(r'LWVR.xlsx')
read_file.to_csv(r"LWVR.csv", index=None, header=True)
os.remove('LWVR.xlsx')

# read source csv file
df = pd.read_csv('LWVR.csv')

# remove rows where Include == 0
df = df[df.Include != 0]

# remove extra columns
df = df.drop('CSA/Planner', axis=1)
df = df.drop('Means of Transport Description', axis=1)
df = df.drop('Distributed to EWM', axis=1)
df = df.drop('Multiple Deliveries on the Shipment', axis=1)
df = df.drop('Resource Max Weight', axis=1)
df = df.drop('Resource Max Weight Status', axis=1)
df = df.drop('Total Weight (Product Dependent Dunnage Calc)', axis=1)
df = df.drop('Available Weight', axis=1)
df = df.drop('Include', axis=1)

# format the data frame
df = df.rename(columns={"Freight Order Number": "Shipment", "Delivery Number": "Delivery",
                        "Planned Load End Wk Nbr": "Plan_WK",
                        "Planned Load End Date": "Plan_Date", "Origin Plant Desc": "Origin",
                        "Dest Location Number": "Destination", "Dest Region": "State",
                        "Transportation Mode Description": "Trans_Mode",
                        "Total Pallets on Freight Order": "Pallets",
                        "Gross Wt LBS (FH)": "Gross_Wt", "To Fill": "To_Fill"})
# create list of Sources
Sources = ["GOLDEN BREWERY", "MILWAUKEE BREWERY", "TRENTON BREWERY",
           "FORT WORTH BREWERY", "SHENANDOAH BREWERY", "ALBANY BREWERY",
           "IRWINDALE BREWERY", "GOLDEN DC", "PORTLAND DC", "ELIZABETH DC",
           "ALBANY DC", "FORT WORTH DC", "MILWAUKEE DC",
           "CHIP FALLS LEINENKUGEL BREWERY", "BACKUS Y JOHNSTON S.A.A.",
           "TYSKIE BROWARY", "BAVARIA S.A.", "GROLSCH BREWERY",
           "BIRRA PERONI Ã‚Â SPA", "CZECH REPUBLIC IMPORT",
           "CANADA,MOLSON, MONTREAL IMPORT", "CANADA, MOLSON, TORONTO IMPORT"]

# replace all unavailable Origins with 0
df.loc[~df["Origin"].isin(Sources), "Origin"] = 0

# create new data frame without Origin == 0
df = df[df.Origin != 0]

# replace column values with replacements
df.loc[df.Origin == "GOLDEN BREWERY", "Origin"] = 1000
df.loc[df.Origin == "MILWAUKEE BREWERY", "Origin"] = 1010
df.loc[df.Origin == "TRENTON BREWERY", "Origin"] = 1020
df.loc[df.Origin == "FORT WORTH BREWERY", "Origin"] = 1030
df.loc[df.Origin == "SHENANDOAH BREWERY", "Origin"] = 1040
df.loc[df.Origin == "ALBANY BREWERY", "Origin"] = 1060
df.loc[df.Origin == "IRWINDALE BREWERY", "Origin"] = 1070
df.loc[df.Origin == "GOLDEN DC", "Origin"] = 2000
df.loc[df.Origin == "PORTLAND DC", "Origin"] = 2020
df.loc[df.Origin == "ELIZABETH DC", "Origin"] = 2030
df.loc[df.Origin == "ALBANY DC", "Origin"] = 2060
df.loc[df.Origin == "FORT WORTH DC", "Origin"] = 2070
df.loc[df.Origin == "MILWAUKEE DC", "Origin"] = 2080
df.loc[df.Origin == "CHIP FALLS LEINENKUGEL BREWERY", "Origin"] = 5000
df.loc[df.Origin == "BACKUS Y JOHNSTON S.A.A.", "Origin"] = 6800
df.loc[df.Origin == "TYSKIE BROWARY", "Origin"] = 6810
df.loc[df.Origin == "BAVARIA S.A.", "Origin"] = 6850
df.loc[df.Origin == "GROLSCH BREWERY", "Origin"] = 6860
df.loc[df.Origin == "BIRRA PERONI Ã‚Â SPA", "Origin"] = 6870
df.loc[df.Origin == "CZECH REPUBLIC IMPORT", "Origin"] = 6880
df.loc[df.Origin == "CANADA,MOLSON, MONTREAL IMPORT", "Origin"] = 6890
df.loc[df.Origin == "CANADA, MOLSON, TORONTO IMPORT", "Origin"] = 6900
df.loc[df.Trans_Mode == "Intermodal", "Trans_Mode"] = "Intm"

# filling missing value using fillna()
df.fillna(0)

# create new columns
# Source ID column -> Source-Trans_Mode-Destination
df['Source_ID'] = df.agg('{0[Origin]}-{0[Trans_Mode]}-{0[Destination]}'.format, axis=1)
# Prefill_%
df['Prefill_%'] = df['Gross_Wt'] / (df['Gross_Wt'] + df['To_Fill'])

# format data frame columns
df = df.astype({"Shipment": int, "Delivery": int, "Plan_WK": int,
                "Origin": int, "Pallets": int, "Gross_Wt": float,
                "To_Fill": float, "Prefill_%": float})

# reindex columns
df = df[['Shipment', 'Delivery', 'Source_ID', 'Plan_WK',
         'Plan_Date', 'Origin', 'Trans_Mode', 'Destination',
         'State', 'Pallets', 'Gross_Wt', 'To_Fill',
         'Prefill_%']]

# convert data frame to dictionary
LWVR = df.set_index('Delivery').T.to_dict('list')

# print execution time
end_time = datetime.datetime.now()
LWVR_Time = end_time - begin_time
print(f"End time:   {end_time}")
print(f"LWVR Time:  {LWVR_Time}")
